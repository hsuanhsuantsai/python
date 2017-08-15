import getpass
import splunklib.client as client
import openpyxl
import sys
import argparse

from openpyxl import Workbook

filename = 'sdkTest.xlsx'
CONNECT_ARGS = {
	'host': 'localhost',
	'port': 8089,
	'username': '',
	'password': '',
    }

# Arg parser
parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', help='Set customized filename')
args = parser.parse_args()
if args.file:
	filename = args.file

try:
	CONNECT_ARGS['username'] = raw_input('Enter splunk username: ')
	CONNECT_ARGS['password'] = getpass.getpass()
except Exception as e:
	print(e)
	sys.exit(-1)

print('Login as : ' + CONNECT_ARGS['username'])
service = client.connect(**CONNECT_ARGS)
savedsearches = service.saved_searches


output_wb = Workbook()
o_sheet = output_wb.create_sheet(title='Sheet1')

o_sheet.cell(row=1, column=1).value = 'Report Title'
o_sheet.cell(row=1, column=2).value = 'Search Query'
o_sheet.cell(row=1, column=3).value = 'Time Range'
o_sheet.cell(row=1, column=4).value = 'Owner'

cur_row = 2
for ss in savedsearches:
	o_sheet.cell(row=cur_row, column=1).value = ss.name
	o_sheet.cell(row=cur_row, column=2).value = ss['search']
	o_sheet.cell(row=cur_row, column=3).value = ss['dispatch.earliest_time']
	o_sheet.cell(row=cur_row, column=4).value = ss.access['owner']
	print(cur_row)
	print(ss.name)
	print(ss['search'])
	print(ss['dispatch.earliest_time'])
	print(ss.access['owner'])
	print('---------------------')
	cur_row += 1



output_wb.save(filename)
print('Saved as ' + filename)


### EOF